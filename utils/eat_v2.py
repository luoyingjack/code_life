import collections
import shutil
from openpyxl import load_workbook
from openpyxl.styles import Font, NamedStyle
from openpyxl.styles import Border, Side, Alignment


address = {
    '6号楼': ['6号楼', 'FM3', 'FM4', 'FM5', '战略发展部', 'PM', 'FM2-6号楼', '销售', 'RA', 'IM', '六号楼', '市场', 'FM36号楼',
            'HR', 'FM3-6号楼', '苗圃6号楼', '战略', 'FM2', '战略6号楼', 'FM26号楼', 'QA', 'FM5-6号楼', '6号楼三楼', '6号楼FM5'],
    '1楼': ['6号楼一楼', '六号楼一楼', '6号楼-1楼', 'FM8-1', 'FM8-1楼', 'FM7-1', '1楼', '一楼', '1号楼'],
    '5号楼': ['5号楼', 'FM1-5号楼', 'FM65号楼', 'FM9-5号楼', 'PM-5号楼', 'FM6-5号楼', 'FM1－5号楼', 'FM15号楼', 'FM95号楼',
            'FM15号楼', 'FM1-5号楼', '五号楼']
}
sort_site = ['6号楼', '1楼', '5号楼']


def tidy_data(data):
    """整理数据"""
    result = []
    staff_list = [line for line in data.split('\n') if line]
    staff_list.pop(0)
    eat_date = staff_list.pop(0)
    for obj in staff_list:
        eat_noon, eat_night, eat_taste = False, False, ''
        obj = obj.replace(' ', '')
        try:
            index, info = obj.split('.')
            if ('中饭' in info) or ('午饭' in info):
                eat_noon = True
                info = info.replace('中饭', '~').replace('午饭', '~')
            if '晚饭' in info:
                eat_night = True
                info = info.replace('晚饭', '~')
            if '不辣' in info:
                eat_taste = '不辣'
                info = info.replace('不辣', '')
            elif '辣' in info:
                eat_taste = '辣'
                info = info.replace('辣', '')
            else:
                eat_taste = '不辣'
            info = info.replace('~~', '~')
            name, branch = info.split('~')
        except Exception as e:
            print(f'[error]={e}, obj={obj}')
        else:
            result.append({
                'index': index,
                'eat_noon': eat_noon,
                'eat_night': eat_night,
                'name': name,
                'branch': branch,
                'eat_taste': eat_taste
            })
    return eat_date, result


def sort_data(data):
    """分类排序"""
    dict_noon, dict_night, noon_names, night_names, total_noon, total_night,  = {}, {}, [], [], 0, 0
    total_noon_taster_hot, total_night_taster_hot = 0, 0
    total_noon_taster_not_hot, total_night_taster_not_hot = 0, 0
    # 统计各部门
    dict_one = {'noon': 0, 'night': 0, 'noon_hot': 0, 'noon_not_hot': 0, 'night_hot': 0, 'night_not_hot': 0}
    dict_five = {'noon': 0, 'night': 0, 'noon_hot': 0, 'noon_not_hot': 0, 'night_hot': 0, 'night_not_hot': 0}
    dict_six = {'noon': 0, 'night': 0, 'noon_hot': 0, 'noon_not_hot': 0, 'night_hot': 0, 'night_not_hot': 0}
    for obj in data:
        index, eat_noon, eat_night, name, branch, eat_taste = map(
            obj.get,
            ['index', 'eat_noon', 'eat_night', 'name', 'branch', 'eat_taste']
        )
        sign, site = False, None
        for k, v in address.items():
            if branch in v:
                site = k
                break
        if not site:
            print(index, branch, '部门未识别')
            continue
        # 午餐统计
        if eat_noon:
            list_noon = dict_noon.get(site)
            if list_noon:
                list_noon.append({'姓名': name, '平台': site, '口味': eat_taste})
            else:
                dict_noon[site] = [{'姓名': name, '平台': site, '口味': eat_taste}]
            noon_names.append(name)
            total_noon += 1
            # 口味统计
            if eat_taste == '辣':
                total_noon_taster_hot += 1
            elif eat_taste == '不辣':
                total_noon_taster_not_hot += 1
            # 数量统计
            if site == '1楼':
                dict_one['noon'] += 1
                if eat_taste == '辣':
                    dict_one['noon_hot'] += 1
                elif eat_taste == '不辣':
                    dict_one['noon_not_hot'] += 1
            elif site == '5号楼':
                dict_five['noon'] += 1
                if eat_taste == '辣':
                    dict_five['noon_hot'] += 1
                elif eat_taste == '不辣':
                    dict_five['noon_not_hot'] += 1
            elif site == '6号楼':
                dict_six['noon'] += 1
                if eat_taste == '辣':
                    dict_six['noon_hot'] += 1
                elif eat_taste == '不辣':
                    dict_six['noon_not_hot'] += 1
        # 晚餐统计
        if eat_night:
            list_night = dict_night.get(site)
            if list_night:
                list_night.append({'姓名': name, '平台': site, '口味': eat_taste})
            else:
                dict_night[site] = [{'姓名': name, '平台': site, '口味': eat_taste}]
            night_names.append(name)
            total_night += 1
            # 口味统计
            if eat_taste == '辣':
                total_night_taster_hot += 1
            elif eat_taste == '不辣':
                total_night_taster_not_hot += 1
            # 数量统计
            if site == '1楼':
                dict_one['night'] += 1
                if eat_taste == '辣':
                    dict_one['night_hot'] += 1
                elif eat_taste == '不辣':
                    dict_one['night_not_hot'] += 1
            elif site == '5号楼':
                dict_five['night'] += 1
                if eat_taste == '辣':
                    dict_five['night_hot'] += 1
                elif eat_taste == '不辣':
                    dict_five['night_not_hot'] += 1
            elif site == '6号楼':
                dict_six['night'] += 1
                if eat_taste == '辣':
                    dict_six['night_hot'] += 1
                elif eat_taste == '不辣':
                    dict_six['night_not_hot'] += 1
    # 检测重复名单
    noon_more = collections.Counter(noon_names)
    night_more = collections.Counter(night_names)
    repetition_noon = [k for k in noon_more if noon_more[k] >= 2]
    repetition_night = [k for k in night_more if night_more[k] >= 2]
    if repetition_noon:
        print('【午饭】请检查疑似重复名单', repetition_noon)
    if repetition_night:
        print('【晚饭】请检查疑似重复名单', repetition_night)
    data = {
        'dict_noon': dict_noon,
        'dict_night': dict_night,
        'total_noon': total_noon,
        'total_night': total_night,
        'total_noon_taster_hot': total_noon_taster_hot,
        'total_noon_taster_not_hot': total_noon_taster_not_hot,
        'total_night_taster_hot': total_night_taster_hot,
        'total_night_taster_not_hot': total_night_taster_not_hot,
        'dict_one': dict_one,
        'dict_five': dict_five,
        'dict_six': dict_six
    }
    return data


def gen_excel(eat_date, data):
    """生成excel表格"""
    # 拷贝文件
    filename = '/Users/jack/Desktop/{0}.xlsx'.format(eat_date)
    shutil.copyfile('../static/订餐模板v2.xlsx', filename)
    # 打开一个将写的文件
    wb = load_workbook(filename)
    sheet = wb['Sheet']
    # 设置格式
    line_t = Side(style='thin', color='000000')  # 细边框
    border = Border(top=line_t, bottom=line_t, left=line_t, right=line_t)
    alignment = Alignment(horizontal='center', vertical='center')
    ft = Font(name='楷体', color='000000', size=48, b=True)
    sty = NamedStyle(name='sty', font=ft, border=border, alignment=alignment)
    #  写入日期
    sheet.cell(column=11, row=1, value=eat_date).style = sty

    # 写入数据
    dict_noon, dict_night, total_noon, total_night, total_noon_taster_hot, total_noon_taster_not_hot, \
    total_night_taster_hot, total_night_taster_not_hot, dict_one, dict_five, dict_six = map(
        data.get,
        ['dict_noon', 'dict_night', 'total_noon', 'total_night', 'total_noon_taster_hot', 'total_noon_taster_not_hot',
         'total_night_taster_hot', 'total_night_taster_not_hot', 'dict_one', 'dict_five', 'dict_six']
    )
    index = 1
    # 写入午餐
    noon_nr, noon_mr = 3, 2
    for site in sort_site:
        site_obj = dict_noon.get(site)
        if not site_obj:
            continue
        noon_mr += 1
        start_row = noon_mr
        m = 0
        for line in site_obj:
            sheet.cell(column=1, row=noon_nr, value=str(index)).style = sty
            sheet.cell(column=2, row=noon_nr, value=line.get('姓名') or '').style = sty
            sheet.cell(column=3, row=noon_nr, value=line.get('平台') or '').style = sty
            sheet.cell(column=4, row=noon_nr, value='午餐').style = sty
            sheet.cell(column=5, row=noon_nr, value=line.get('口味') or '').style = sty
            noon_nr += 1
            noon_mr += 1
            index += 1
            m += 1
        noon_mr -= 1
        stop_row = noon_mr
        sheet.merge_cells(start_row=start_row, end_row=stop_row, start_column=6,  end_column=6)
        sheet.cell(column=6, row=start_row, value=m).style = sty

    # 写入晚餐
    night_nr, night_mr = 3, 2
    for site in sort_site:
        site_obj = dict_night.get(site)
        if not site_obj:
            continue
        night_mr += 1
        start_row = night_mr
        x = 0
        for line in site_obj:
            sheet.cell(column=7, row=night_nr, value=line.get('姓名') or '').style = sty
            sheet.cell(column=8, row=night_nr, value=line.get('平台') or '').style = sty
            sheet.cell(column=9, row=night_nr, value='晚餐').style = sty
            sheet.cell(column=10, row=night_nr, value=line.get('口味') or '').style = sty
            night_nr += 1
            night_mr += 1
            x += 1
        night_mr -= 1
        stop_row = night_mr
        sheet.merge_cells(start_row=start_row, end_row=stop_row, start_column=11, end_column=11)
        sheet.cell(column=11, row=start_row, value=x).style = sty

    sheet.cell(column=1, row=noon_nr, value='合计').style = sty
    sheet.merge_cells(start_row=noon_nr, end_row=noon_nr, start_column=2, end_column=5)
    sheet.merge_cells(start_row=noon_nr + 1, end_row=noon_nr + 1, start_column=2, end_column=5)
    sheet.merge_cells(start_row=noon_nr + 2, end_row=noon_nr + 2, start_column=2, end_column=5)
    sheet.merge_cells(start_row=noon_nr + 3, end_row=noon_nr + 3, start_column=2, end_column=5)
    sheet.merge_cells(start_row=noon_nr + 4, end_row=noon_nr + 4, start_column=2, end_column=5)
    sheet.merge_cells(start_row=noon_nr + 5, end_row=noon_nr + 5, start_column=2, end_column=5)
    sheet.merge_cells(start_row=noon_nr + 6, end_row=noon_nr + 6, start_column=2, end_column=5)
    sheet.merge_cells(start_row=noon_nr + 7, end_row=noon_nr + 7, start_column=2, end_column=5)

    sheet.cell(column=2, row=noon_nr, value=f'午饭共{total_noon}份, 有{total_noon_taster_hot}份辣， {total_noon_taster_not_hot}份不辣').style = sty
    sheet.cell(column=2, row=noon_nr + 1, value=f'晚饭共{total_night}份，有{total_night_taster_hot}份辣， {total_night_taster_not_hot}份不辣').style = sty
    sheet.cell(column=2, row=noon_nr + 2, value=f'1楼午饭共{dict_one["noon"]}份, 有{dict_one["noon_hot"]}份辣， {dict_one["noon_not_hot"]}份不辣').style = sty
    sheet.cell(column=2, row=noon_nr + 3, value=f'1楼晚饭共{dict_one["night"]}份, 有{dict_one["night_hot"]}份辣， {dict_one["night_not_hot"]}份不辣').style = sty
    sheet.cell(column=2, row=noon_nr + 4, value=f'5号楼午饭共{dict_five["noon"]}份, 有{dict_five["noon_hot"]}份辣， {dict_five["noon_not_hot"]}份不辣').style = sty
    sheet.cell(column=2, row=noon_nr + 5, value=f'5号楼晚饭共{dict_five["night"]}份, 有{dict_five["night_hot"]}份辣， {dict_five["night_not_hot"]}份不辣').style = sty
    sheet.cell(column=2, row=noon_nr + 6, value=f'6号楼午饭共{dict_six["noon"]}份, 有{dict_six["noon_hot"]}份辣， {dict_six["noon_not_hot"]}份不辣').style = sty
    sheet.cell(column=2, row=noon_nr + 7, value=f'6号楼晚饭共{dict_six["night"]}份, 有{dict_six["night_hot"]}份辣， {dict_six["night_not_hot"]}份不辣').style = sty

    wb.save(filename)
    return 'success'