import io
import os
import shutil
import ssl
import urllib.request
import zipfile

import xlsxwriter
from app.models import UserDemo, Authorizer
# 跳过ssl证书验证
from utils.string_util import gen_random_str

context = ssl._create_unverified_context()


def user_data():
    """
    导出excel中插入图片
    :return:
    """
    query = UserDemo.select()
    book = xlsxwriter.Workbook('用户表.xlsx')
    sheet = book.add_worksheet()
    titles = ['昵称', '头像']
    for d in range(len(titles)):
        sheet.write(0, d, titles[d])
    n = 1
    for user in query.order_by(UserDemo.id.asc()):
        info = [user.nickname]
        for d in range(len(info)):
            sheet.write(n, d, info[d])
        # 插入图片
        image_data = io.BytesIO(urllib.request.urlopen(user.headimgurl, context=context).read())
        sheet.insert_image(n, 1, user.nickname, options={'image_data': image_data, 'x_scale': 0.2, 'y_scale': 0.2})
        n += 1
    book.close()
    return 'success'


'''
Description： 读取excel中的图片，打印图片路径
  先将excel转换成zip包，解压zip包，包下面有文件夹存放了图片，读取这个图片
'''


def isfile_exist(file_path):
    """判断是否是文件和判断文件是否存在"""
    if not os.path.isfile(file_path):
        print("It's not a file or no such file exist ! %s" % file_path)
        return False
    else:
        return True


def change_file_name(file_path, new_type='.zip'):
    """修改指定目录下的文件类型名，将excel后缀名修改为.zip"""
    if not isfile_exist(file_path):
        return ''
    extend = os.path.splitext(file_path)[1]  # 获取文件拓展名
    if extend != '.xlsx' and extend != '.xls':
        print("It's not a excel file! %s" % file_path)
        return False
    file_name = os.path.basename(file_path)  # 获取文件名
    new_name = str(file_name.split('.')[0]) + new_type  # 新的文件名，命名为：xxx.zip
    dir_path = os.path.dirname(file_path)  # 获取文件所在目录
    new_path = os.path.join(dir_path, new_name)  # 新的文件路径
    if os.path.exists(new_path):
        os.remove(new_path)
    shutil.copy(os.path.join(file_path), '/Users/jack/Desktop/test.xlsx')  # 复制出来一份新文件
    os.rename(file_path, new_path)  # 保存新文件，旧文件会替换掉
    return new_path  # 返回新的文件路径，压缩包


def unzip_file(zipfile_path):  # 解压文件
    if not isfile_exist(zipfile_path):
        return False
    if os.path.splitext(zipfile_path)[1] != '.zip':
        print("It's not a zip file! %s" % zipfile_path)
        return False
    file_zip = zipfile.ZipFile(zipfile_path, 'r')
    file_name = os.path.basename(zipfile_path)  # 获取文件名
    zipdir = os.path.join(os.path.dirname(zipfile_path), str(file_name.split('.')[0]))  # 获取文件所在目录
    for files in file_zip.namelist():
        file_zip.extract(files, os.path.join(zipfile_path, zipdir))  # 解压到指定文件目录
    file_zip.close()
    return True


def read_img(zipfile_path):
    """读取解压后的文件夹，打印图片路径"""
    if not isfile_exist(zipfile_path):
        return False
    dir_path = os.path.dirname(zipfile_path)  # 获取文件所在目录
    file_name = os.path.basename(zipfile_path)  # 获取文件名
    pic_dir = 'xl' + os.sep + 'media'  # excel变成压缩包后，再解压，图片在media目录
    pic_path = os.path.join(dir_path, str(file_name.split('.')[0]), pic_dir)
    file_list = os.listdir(pic_path)
    for file in file_list:
        filepath = os.path.join(pic_path, file)
        print(filepath)


def compenent(excel_file_path):
    """组合各个函数"""
    zip_file_path = change_file_name(excel_file_path)
    if zip_file_path != '':
        if unzip_file(zip_file_path):
            read_img(zip_file_path)


def export_platform_user_openid(a_id: int):
    """导出公众号粉丝openid上传七牛"""
    import openpyxl
    from utils.service_util import qn_service
    from io import BytesIO
    authorizer = Authorizer.get_by_id(a_id)
    book = openpyxl.Workbook()
    sheet = book.create_sheet(index=0)
    i = 1
    total, all_count = 1, 0
    next_openid = ''
    while total > all_count:
        try:
            ret = authorizer.get_users(next_openid)
            if ret.get('errcode'):
                print(f'Ignore: 授权更新粉丝头像，platform = {authorizer.id}, {ret=}')
                break
            data, total, count, next_openid = map(ret.get, ['data', 'total', 'count', 'next_openid'])
            if data:
                openid_list = data.get('openid')
            else:
                openid_list = []
            all_count += count
        except Exception as e:
            print(f'Fail: 授权更新粉丝头像，platform = {authorizer.id}', e)
            break

        for openid in openid_list:
            sheet.cell(column=1, row=i, value=openid)
            if i and i % 10000 == 0:
                print(f'运行数量={i}', flush=True)
            i += 1
    key = gen_random_str(20, prefix='tmp/', suffix='.xls')
    output = BytesIO()
    book.save(output)
    output.seek(0)
    url = qn_service.upload_data(key, output.getvalue())
    print(f'url={url}', flush=True)


if __name__ == '__main__':
    compenent('/Users/jack/Desktop/test.xlsx')
